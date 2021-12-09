import sqlite3
import sys
from time import sleep

import openpyxl
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QPixmap
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QInputDialog, QLabel, QMessageBox, QFileDialog

from additional_windows import Errors, StarWindow
from interfeic6 import Ui_MainWindow


class MyWidget(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.startWindow = StarWindow()
        sleep(3)
        self.startWindow.close()
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        self.list_of_tabs.setCurrentIndex(0)
        self.NameSchool.setAlignment(Qt.AlignCenter)
        self.dic_time_management = {}
        self.dic_id_subjects = {}
        subjects = cur.execute("""SELECT * FROM subjects_in_school""").fetchall()
        for i in subjects:
            self.dic_id_subjects[i[0]] = i[1]
        self.table_of_classes.setColumnCount(3)
        self.filling_in_the_settings()
        self.filling_in_the_teachers()
        self.filling_in_subjects()
        self.filling_loads()
        self.filling_in_classes()
        self.filling_out_the_schedule()
        self.table_of_teachers.setColumnHidden(3, True)
        self.table_of_classes.setColumnHidden(2, True)
        self.table_of_subjects.setColumnHidden(2, True)
        self.table_of_loads.setColumnHidden(4, True)
        self.g = False
        con.close()
        self.connectiing_all()

    def filling_in_the_teachers(self):
        # Начальная загрузка из базы данных и заполнение таблицы учителей
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        name = cur.execute("""SELECT * FROM list_of_teachers""").fetchall()
        for i in range(len(name)):
            self.table_of_teachers.insertRow(self.table_of_teachers.rowCount())
            item_of_name = QTableWidgetItem(name[i][1])
            item_of_surname = QTableWidgetItem(name[i][2])
            item_of_MiddleName = QTableWidgetItem(name[i][3])
            item_id = QTableWidgetItem(str(name[i][0]))
            self.table_of_teachers.setItem(i, 0, item_of_name)
            self.table_of_teachers.setItem(i, 2, item_of_MiddleName)
            self.table_of_teachers.setItem(i, 1, item_of_surname)
            self.table_of_teachers.setItem(i, 3, item_id)

    def filling_in_the_settings(self):
        settings = open("txt files/settings.txt", "r").read().split("\n")
        self.name_of_school.setText(settings[0])
        self.NameSchool.setText(f"Расписание школы {settings[0]}")
        for i in range(1, 7):
            self.dic_time_management[i] = settings[i]
        # Заполнение Настроек из settings.txt
        self.dic_time_management[0] = settings[6]
        self.one_lesson.setText(settings[1])
        self.two_lesson.setText(settings[2])
        self.three_lesson.setText(settings[3])
        self.fout_lesson.setText(settings[4])
        self.five_lesson.setText(settings[5])
        self.six_lesson.setText(settings[6])
        self.seven_lesson.setText(settings[7])
        self.eight_lesson.setText(settings[8])
        # Устанавливение картинок: "YandexEmblem.png", "emblem.gif" из папки images
        pixmap = QPixmap("Images/YandexEmblema.png")
        self.emblema_yandex.setPixmap(pixmap)
        pixmap = QPixmap("Images/emblem.gif")
        self.emblema.setPixmap(pixmap)
        self.list__of_classes = []

    def filling_in_subjects(self):
        self.subjects = []
        # Начальная загрузка из базы данных и заполнение таблицы предметов
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        name = cur.execute("""SELECT * FROM subjects_in_school""").fetchall()[1:]
        for i in range(len(name)):
            self.table_of_subjects.insertRow(self.table_of_subjects.rowCount())
            item_of_name_subject = QTableWidgetItem(name[i][1])
            item_of_load = QTableWidgetItem(str(name[i][2]))
            item_id = QTableWidgetItem(str(name[i][0]))
            self.subjects.append(item_of_name_subject.text())
            self.table_of_subjects.setItem(i, 0, item_of_name_subject)
            self.table_of_subjects.setItem(i, 1, item_of_load)
            self.table_of_subjects.setItem(i, 2, item_id)

    def filling_in_classes(self):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        # Начальная загрузка из базы данных и заполнение таблицы классов
        name = cur.execute("""SELECT * FROM list_of_classes""").fetchall()
        for i in range(len(name)):
            self.table_of_classes.insertRow(self.table_of_classes.rowCount())
            item_of_name_class = QTableWidgetItem(name[i][1])
            item_id = QTableWidgetItem(str(name[i][0]))
            self.list__of_classes.append(item_of_name_class.text())
            self.table_of_classes.setItem(i, 0, item_of_name_class)
            self.table_of_classes.setItem(i, 2, item_id)

    def connectiing_all(self):
        # Подключение кнопок к функциям
        self.table_of_subjects.cellChanged.connect(self.call_update_subjects)
        self.table_of_classes.cellChanged.connect(self.call_update_classes)
        self.table_of_teachers.cellChanged.connect(self.call_update_teachers)
        self.table_of_timetable.cellClicked.connect(self.update_timetable)
        self.btn_check.clicked.connect(self.check_timetable)
        self.delete_btn_loads.clicked.connect(self.delete_loads)
        self.add_btn_teachers.clicked.connect(lambda: self.add(self.table_of_teachers, "list_of_teachers"))
        self.add_btn_school_objects.clicked.connect(lambda: self.add(self.table_of_subjects, "subjects_in_school"))
        self.add_btn_school_classes.clicked.connect(self.add_class)
        self.add_btn_loads.clicked.connect(self.add_loads)
        self.delete_btn_teachers.clicked.connect(lambda: self.delete(self.table_of_teachers, "list_of_teachers"))
        self.delete_btn_school_classes.clicked.connect(lambda: self.delete(self.table_of_classes, "list_of_classes"))
        self.delete_btn_school_objects.clicked.connect(
            lambda: self.delete(self.table_of_subjects, "subjects_in_school"))
        self.save_settings.clicked.connect(self.saving__settings)
        self.save_name_of_school.clicked.connect(self.saving_name_of_school)
        self.btn_timetable_to_excel.clicked.connect(self.timetable_to_excel)
        self.filter_for_loads_on_class.currentTextChanged.connect(self.filter)
        self.table_of_loads.cellClicked.connect(self.update_loads)

    def filling_loads(self):  # Загрузка данных из BD в таблицу 'нагрузки'
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        name = cur.execute(f"""SELECT * FROM list_of_classes""").fetchall()
        loads = cur.execute('SELECT * FROM list_of_loads').fetchall()  # Получаем из базы данных нагрузки
        len_loads = len(loads)  # Запоминаем длину
        self.table_of_loads.setRowCount(len_loads)  # Заполняем таблицу нагрузок
        teachers = cur.execute("""SELECT * FROM list_of_teachers""").fetchall()  # Получаем учителей из базы данных
        self.class_dic_id = {}  # Создаём словарь где будет хранится id класса и его название например 5: "1 'А' класс"
        for i in name:
            self.class_dic_id[i[0]] = i[1]  # Заполняем словарь
        self.dic_teachers = {}  # Создаём словарь где будет хранится id учителя
        # и его имя например 5: "Макаренко Мария Дмитриевна"
        for i in teachers:
            i = list(i)
            if not i[1]:
                i[1] = ""
            if not i[3]:
                i[3] = ""
            if not i[2]:
                i[2] = ""
            self.dic_teachers[i[0]] = i[2] + " " + i[1] + " " + i[3]  # Заполняем словарь

        # Заполняем Таблицу нагрузок виджетами QLabel
        for i in range(len_loads):
            name_of_teacher = QLabel(self)
            if loads[i][0]:
                name_of_teacher.setText(str(self.dic_teachers[int(loads[i][0])]))
            else:
                name_of_teacher.setText("")
            load = QLabel(self)
            if loads[i][3]:
                load.setText(str(int(loads[i][3])))
            else:
                load.setText("")
            load.setAlignment(Qt.AlignHCenter)
            load.setText(str(int(loads[i][3])))
            class_name = QLabel(self)
            if loads[i][1]:
                class_name.setText(self.class_dic_id[int(loads[i][1])])
            else:
                class_name.setText("")
            subject = QLabel(self)
            if loads[i][2]:
                subject.setText(str(self.dic_id_subjects[int(loads[i][2])]))
            else:
                subject.setText("")
            id = QLabel(self)
            id.setText(str(int(loads[i][4])))
            self.table_of_loads.setCellWidget(i, 0, name_of_teacher)
            self.table_of_loads.setCellWidget(i, 3, load)
            self.table_of_loads.setCellWidget(i, 1, class_name)
            self.table_of_loads.setCellWidget(i, 2, subject)
            self.table_of_loads.setCellWidget(i, 4, id)  # fake3
            self.paste_to_table_of_loads(i, name_of_teacher, load, class_name, subject, id)
        self.table_of_loads.setColumnWidth(0, 250)

    def paste_to_table_of_loads(self, i, name_of_teacher, load, class_name, subject, id):
        self.table_of_loads.setCellWidget(i, 0, name_of_teacher)
        self.table_of_loads.setCellWidget(i, 3, load)
        self.table_of_loads.setCellWidget(i, 1, class_name)
        self.table_of_loads.setCellWidget(i, 2, subject)
        self.table_of_loads.setCellWidget(i, 4, id)

    # Заполнение расписания
    def filling_out_the_schedule(self):
        # Создание таблицы Расписания
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        self.table_of_timetable.insertColumn(self.table_of_timetable.columnCount())
        item = QTableWidgetItem("")
        self.table_of_timetable.setHorizontalHeaderItem(self.table_of_timetable.columnCount() - 1, item)
        self.table_of_timetable.insertColumn(self.table_of_timetable.columnCount())
        self.table_of_timetable.setHorizontalHeaderItem(self.table_of_timetable.columnCount() - 1, item)
        self.table_of_timetable.insertColumn(self.table_of_timetable.columnCount())
        self.table_of_timetable.setHorizontalHeaderItem(self.table_of_timetable.columnCount() - 1, item)
        self.list_of_tabs.currentChanged.connect(self.update_tabs)
        name = cur.execute("""SELECT * FROM list_of_classes""").fetchall()
        # Создание столбцов Расписания
        name = list(map(list, name))
        for i in range(len(self.list__of_classes)):
            self.table_of_timetable.insertColumn(self.table_of_timetable.columnCount())
            item = QTableWidgetItem(self.list__of_classes[i])
            item.setFont(QFont("Impact", 14))
            self.table_of_timetable.setColumnWidth(self.table_of_timetable.columnCount() - 1, 200)
            self.table_of_timetable.setHorizontalHeaderItem(self.table_of_timetable.columnCount() - 1, item)
        self.table_of_timetable.setRowCount(36)

        # Заполнение строк Расписания
        k = 1
        result = cur.execute("""SELECT * FROM timetable""").fetchall()
        dic_indexes_columns_classes = {}
        for i in name:
            dic_indexes_columns_classes[i[1]] = 3 + name.index(i)
        for i in range(36):
            if k > 6:
                k = 1
            item = (QTableWidgetItem(str(k)))
            item.setTextAlignment(Qt.AlignCenter)
            self.table_of_timetable.setItem(i, 1, item)
            self.table_of_timetable.setItem(i, 2, QTableWidgetItem(self.dic_time_management[k]))
            k += 1
        for i in range(len(result)):
            text = self.dic_id_subjects[result[i][3]]
            image = QLabel(text, self)
            image.setFont(QFont("Times New Roman", 13))
            image.setAlignment(Qt.AlignCenter)
            day = result[i][1]
            number_lesson = result[i][2]
            self.table_of_timetable.setCellWidget((day - 1) * 6 + number_lesson - 1,
                                                  dic_indexes_columns_classes[self.class_dic_id[result[i][0]]], image)
        for i in range(36):
            for j in range(3, self.table_of_timetable.columnCount()):
                item = self.table_of_timetable.cellWidget(i, j)
                if item is None:
                    image = QLabel("", self)
                    image.setFont(QFont("Times New Roman", 13))
                    image.setAlignment(Qt.AlignCenter)
                    self.table_of_timetable.setCellWidget(i, j, image)

        # Создаём заголовки слева для вклаки расписание
        for i in range(0, 37, 6):
            self.table_of_timetable.setSpan(i, 0, 6, 1)
        list_ = ['П\nо\nн\nе\nд\nе\nл\nь\nн\nи\nк',
                  'В\nт\nо\nр\nн\nи\nк',
                  'С\nр\nе\nд\nа',
                  'Ч\nе\nт\nв\nе\nр\nг',
                  'П\nя\nт\nн\nи\nц\nа',
                  'С\nу\nб\nб\nо\nт\nа']

        for a in range(6):
            item = QTableWidgetItem(list_[a])
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Arial Black", 8))
            self.table_of_timetable.setItem(a * 6, 0, item)
        self.table_of_timetable.setColumnWidth(0, 5)
        self.table_of_timetable.setColumnWidth(1, 1)
        self.table_of_timetable.setColumnWidth(2, 70)
        self.table_of_timetable.setItem(6, 0, item)

    def nobody(self):
        pass

    def filter(self, value):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        name = cur.execute("""SELECT * FROM list_of_classes""").fetchall()
        if value != "Все":
            self.table_of_loads.setRowCount(0)
            loads = cur.execute(
                f"""SELECT * FROM list_of_loads 
                WHERE name_class in (SELECT id FROM list_of_classes 
                WHERE name_class = '{value}')""").fetchall()  # Получаем из базы данных нагрузки
            len_loads = len(loads)  # Запоминаем длину
            self.table_of_loads.setRowCount(len_loads)  # Заполняем таблицу нагрузок
            teachers = cur.execute("""SELECT * FROM list_of_teachers""").fetchall()  # Получаем учителей из базы данных
            self.class_dic_id = {}  # Создаём словарь где будет хранится
            # id класса и его название например 5: "1 'А' класс"
            for i in name:
                self.class_dic_id[i[0]] = i[1]  # Заполняем словарь
            self.dic_teachers = {}  # Создаём словарь где будет хранится
            # id учителя и его имя например 5: "Макаренко Мария Дмитриевна"
            for i in teachers:
                self.dic_teachers[i[0]] = i[2] + " " + i[1] + " " + i[3]  # Заполняем словарь

            # Заполняем Таблицу нагрузок виджетами QLabel
            for i in range(len_loads):
                name_of_teacher = QLabel(self)
                if loads[i][0]:
                    name_of_teacher.setText(str(self.dic_teachers[int(loads[i][0])]))
                else:
                    name_of_teacher.setText("")
                load = QLabel(self)
                if loads[i][3]:
                    load.setText(str(int(loads[i][3])))
                else:
                    load.setText("")
                load.setAlignment(Qt.AlignHCenter)
                load.setText(str(int(loads[i][3])))
                class_name = QLabel(self)
                if loads[i][1]:
                    class_name.setText(self.class_dic_id[int(loads[i][1])])
                else:
                    class_name.setText("")
                subject = QLabel(self)
                if loads[i][2]:
                    subject.setText(str(self.dic_id_subjects[int(loads[i][2])]))
                else:
                    subject.setText("")
                id = QLabel(self)
                id.setText(str(int(loads[i][4])))
                self.table_of_loads.setCellWidget(i, 0, name_of_teacher)
                self.table_of_loads.setCellWidget(i, 3, load)
                self.table_of_loads.setCellWidget(i, 1, class_name)
                self.table_of_loads.setCellWidget(i, 2, subject)
                self.table_of_loads.setCellWidget(i, 4, id)  # fake4
                self.paste_to_table_of_loads(i, name_of_teacher, load, class_name, subject, id)
            con.close()
        else:
            # Загрузка данных из BD в таблицу 'нагрузки'
            loads = cur.execute('SELECT * FROM list_of_loads').fetchall()  # Получаем из базы данных нагрузки
            len_loads = len(loads)  # Запоминаем длину
            self.table_of_loads.setRowCount(len_loads)  # Заполняем таблицу нагрузок
            teachers = cur.execute("""SELECT * FROM list_of_teachers""").fetchall()  # Получаем учителей из базы данных
            self.class_dic_id = {}  # Создаём словарь где будет хранится
            # id класса и его название например 5: "1 'А' класс"
            for i in name:
                self.class_dic_id[i[0]] = i[1]  # Заполняем словарь
            self.dic_teachers = {}  # Создаём словарь где будет хранится
            # id учителя и его имя например 5: "Макаренко Мария Дмитриевна"
            for i in teachers:
                i = list(i)
                if not i[1]:
                    i[1] = ""
                if not i[3]:
                    i[3] = ""
                if not i[2]:
                    i[2] = ""
                self.dic_teachers[i[0]] = i[2] + " " + i[1] + " " + i[3]  # Заполняем словарь

            # Заполняем Таблицу нагрузок виджетами QLabel
            for i in range(len_loads):
                name_of_teacher = QLabel(self)
                if loads[i][0]:
                    name_of_teacher.setText(str(self.dic_teachers[int(loads[i][0])]))
                else:
                    name_of_teacher.setText("")
                load = QLabel(self)
                if loads[i][3]:
                    load.setText(str(int(loads[i][3])))
                else:
                    load.setText("")
                load.setAlignment(Qt.AlignHCenter)
                load.setText(str(int(loads[i][3])))
                class_name = QLabel(self)
                if loads[i][1]:
                    class_name.setText(self.class_dic_id[int(loads[i][1])])
                else:
                    class_name.setText("")
                subject = QLabel(self)
                if loads[i][2]:
                    subject.setText(str(self.dic_id_subjects[int(loads[i][2])]))
                else:
                    subject.setText("")
                id = QLabel(self)
                id.setText(str(int(loads[i][4])))
                self.table_of_loads.setCellWidget(i, 0, name_of_teacher)
                self.table_of_loads.setCellWidget(i, 3, load)
                self.table_of_loads.setCellWidget(i, 1, class_name)
                self.table_of_loads.setCellWidget(i, 2, subject)
                self.table_of_loads.setCellWidget(i, 4, id)

    # Выгрузка расписания в Excel файл
    def timetable_to_excel(self):
        # Получаем путь и наименование файла через диалоговое окно который нужно сохранить
        fname, answer = QFileDialog.getSaveFileName(self, 'Выбрать куда сохранить', '',
                                                    'Таблица Excel (*.xlsx);;')
        if answer:
            con = sqlite3.connect("BD/base_school.db")
            writer = openpyxl.load_workbook('templates/template.xlsx')  # открываем шаблон Excel для работы
            excel_sheet = writer['Расписание']  # Находим вкладку шаблона "Расписание"
            for i in range(3, self.table_of_timetable.columnCount()):
                for j in range(36):
                    excel_sheet.cell(row=j + 3, column=i + 1).value = self.table_of_timetable.cellWidget(j, i).text()
            for i in range(1, 3):
                for j in range(36):
                    excel_sheet.cell(row=j + 3, column=i + 1).value = self.table_of_timetable.item(j, i).text()
            for i in range(3, self.table_of_timetable.columnCount()):
                excel_sheet.cell(row=2, column=i + 1).value = self.table_of_timetable.horizontalHeaderItem(i).text()
            writer.save(fname)  # Сохраняем в новый выбраный путь и файл пользователем
            con.close()

    # Обновление вкладки на которую перешёл
    def update_tabs(self, value):
        if value == 4:
            self.filter_for_loads_on_class.currentTextChanged.connect(self.nobody)
            self.filter_for_loads_on_class.clear()
            # Все - это значит выбраны все классы на фильтре
            self.filter_for_loads_on_class.addItem("Все")
            con = sqlite3.connect("BD/base_school.db")
            cur = con.cursor()
            self.filter_for_loads_on_class.addItems(list(map(lambda x: x[0],
                                                             cur.execute(f"""SELECT 
                                                      name_class 
                                                      FROM list_of_classes""").fetchall())))
            con.close()
            self.filter_for_loads_on_class.currentTextChanged.connect(self.filter)
        if value == 0:
            self.NameSchool.setText(f"Расписание школы {self.name_of_school.text()}")
            self.NameSchool.setAlignment(Qt.AlignCenter)
            con = sqlite3.connect("BD/base_school.db")
            cur = con.cursor()
            name = []
            index = 0
            classes = list(map(lambda x: x[0], cur.execute("""SELECT name_class FROM list_of_classes"""))) + [""]
            for i in range(self.table_of_timetable.columnCount()):
                text = self.table_of_timetable.horizontalHeaderItem(index).text()
                if text not in classes:
                    self.table_of_timetable.removeColumn(index)
                    index -= 1
                name.append(text)
                index += 1
            for i in range(len(classes)):
                if not classes[i]:
                    classes[i] = ""
                if classes[i] not in name:
                    self.table_of_timetable.insertColumn(self.table_of_timetable.columnCount())
                    item = QTableWidgetItem(classes[i])
                    item.setFont(QFont("Impact", 14))
                    self.table_of_timetable.setColumnWidth(self.table_of_timetable.columnCount() - 1, 200)
                    self.table_of_timetable.setHorizontalHeaderItem(self.table_of_timetable.columnCount() - 1, item)
                    for i in range(36):
                        item = QLabel(self)
                        item.setText("")
                        self.table_of_timetable.setCellWidget(i, self.table_of_timetable.columnCount() - 1, item)
            con.close()

    def check_timetable(self):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        result = cur.execute("""SELECT * FROM timetable""").fetchall()
        h = []
        teachers = []
        classes = []
        for i in range(len(result)):
            k = 0
            id_teacher = result[i][-2]
            id_class = result[i][0]
            teachers.append(id_teacher)
            classes.append(id_class)
            id_subject = result[i][-3]
            for j in range(len(result)):
                if j != i:
                    if result[j][-2] == id_teacher \
                            and result[j][0] == id_class \
                            and result[j][-3] == id_subject:
                        k += 1
            verni_k = cur.execute(f"""SELECT load FROM list_of_loads 
            WHERE name_teacher = {id_teacher} 
            and name_class = {id_class} 
            and subject = {id_subject}""").fetchall()[0][0]
            subject = cur.execute(f"""SELECT Name FROM subjects_in_school WHERE id = {id_subject}""").fetchall()[0][0]
            clas = cur.execute(f"""SELECT name_class FROM list_of_classes WHERE id = {id_class}""").fetchall()[0][0]
            teacher = " ".join(list(cur.execute(
                f"""SELECT Surname, Name, Middle_name FROM list_of_teachers WHERE id = {id_teacher}""").fetchall()[0]))
            if [0, teacher, clas, subject, verni_k, k, abs(verni_k - k)] not in h:
                h.append([0, teacher, clas, subject, verni_k, abs(verni_k - k), k])
        teachers = list(set(teachers))
        for i in range(len(teachers)):
            for j in range(1, 7):
                for g in range(1, 7):
                    teacher = " ".join(list(cur.execute(
                        f"""SELECT Surname, Name, Middle_name FROM list_of_teachers 
                        WHERE id = {teachers[i]}""").fetchall()[0]))
                    answer = cur.execute(f"""SELECT name_class FROM list_of_classes 
                    WHERE id in (SELECT id_class FROM timetable 
                    WHERE day_of_week = {j} and number_lesson = {g} and id_teacher = {teachers[i]})""").fetchall()
                    if len(answer) >= 2:
                        h.append([1, teacher, j, g, answer])
        self.window = Errors(h)
        con.close()

    def add_loads(self):
        self.table_of_loads.insertRow(self.table_of_loads.rowCount())
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        name_of_teacher = QLabel(self)
        name_of_teacher.setText("")
        load = QLabel(self)
        load.setText("0")
        class_name = QLabel(self)
        class_name.setText("")
        subject = QLabel(self)
        subject.setText("")
        id = QLabel(self)
        id.setText("")
        self.table_of_loads.setCellWidget(self.table_of_loads.rowCount() - 1, 0, name_of_teacher)
        self.table_of_loads.setCellWidget(self.table_of_loads.rowCount() - 1, 3, load)
        self.table_of_loads.setCellWidget(self.table_of_loads.rowCount() - 1, 1, class_name)
        self.table_of_loads.setCellWidget(self.table_of_loads.rowCount() - 1, 2, subject)
        cur.execute(f"""INSERT INTO list_of_loads 
        VALUES ({("'', " * 3) + "0" + ", NULL"})""")
        comanda = "SELECT * FROM list_of_loads"
        result = cur.execute(comanda).fetchall()
        name = []
        for i in result:
            if set(i[:-1]) == {"", 0}:
                name.append(i[-1])
        id_ = name[-1]
        id = QLabel(self)
        id.setText(str(id_))
        self.table_of_loads.setCellWidget(self.table_of_loads.rowCount() - 1, 4, id)
        print(id.text(), self.table_of_loads.cellWidget(self.table_of_loads.rowCount() - 1, 4).text())
        con.commit()
        con.close()

    def delete_loads(self):
        valid = QMessageBox.question(
            self, 'Удаление', 'Действительно хотите удалить этот элемент',
            QMessageBox.Yes, QMessageBox.No)
        if valid == QMessageBox.Yes:
            con = sqlite3.connect("BD/base_school.db")
            cur = con.cursor()
            self.g = False
            current = self.table_of_loads.currentRow()
            id = int(self.table_of_loads.cellWidget(current, 4).text())
            result = cur.execute(
                f"""SELECT day_of_week, number_lesson, id_class FROM timetable
                 WHERE id_class in (SELECT name_class FROM list_of_loads 
                 WHERE id = {id}) and id_subject in (SELECT subject FROM list_of_loads 
                 WHERE id = {id}) and id_teacher in (SELECT name_teacher FROM list_of_loads 
                 WHERE id = {id})""").fetchall()
            print(f"""SELECT day_of_week, number_lesson, id_class FROM timetable
                 WHERE id_class in (SELECT name_class FROM list_of_loads 
                 WHERE id = {id}) and id_subject in (SELECT subject FROM list_of_loads 
                 WHERE id = {id}) and id_teacher in (SELECT name_teacher FROM list_of_loads 
                 WHERE id = {id})""")
            if result:
                valid = QMessageBox.question(
                    self, 'Предупреждение', 'Эта нагрузка используется в расписании. '
                                            'Точно хотите удалить этот элемент вместе '
                                            'с ним удалится все уэлементы в расписании'
                                            ' связанные с этой нагрузкой',
                    QMessageBox.Yes, QMessageBox.No)
                dic = {}
                for i in range(3, self.table_of_timetable.columnCount()):
                    dic[self.table_of_timetable.horizontalHeaderItem(i).text()] = i
                if valid == QMessageBox.Yes:
                    cur.execute(
                        f"""DELETE FROM timetable 
                        WHERE id_class in (SELECT name_class FROM list_of_loads 
                        WHERE id = {id}) and id_subject in (SELECT subject FROM list_of_loads 
                        WHERE id = {id}) and id_teacher in (SELECT name_teacher FROM list_of_loads 
                        WHERE id = {id})""")
                    self.table_of_loads.removeRow(current)
                    cur.execute(f"""DELETE FROM list_of_loads
                    WHERE id = {id}""")
                    dic_index_column_class_name_in_timetable = {}
                    classes = cur.execute(f"""SELECT * FROM list_of_classes""").fetchall()
                    for i in range(len(classes)):
                        dic_index_column_class_name_in_timetable[classes[i][0]] = i + 3
                    for i in result:
                        self.table_of_timetable.cellWidget((i[0] - 1) * 6 + i[1] - 1,
                                                           dic_index_column_class_name_in_timetable[i[2]]).setText("")
                    con.commit()
                else:
                    pass
            else:
                current = self.table_of_loads.currentRow()
                id = int(self.table_of_loads.cellWidget(current, 4).text())
                self.table_of_loads.removeRow(current)
                cur.execute(f"""DELETE FROM list_of_loads
                WHERE id = {id}""")
                con.commit()
            con.close()
        else:
            pass

    def add_class(self):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        self.table_of_classes.insertRow(self.table_of_classes.rowCount())
        cur.execute(f"""INSERT INTO list_of_classes 
        VALUES (NULL, NULL)""")

        comanda = "SELECT * FROM list_of_classes"
        result = cur.execute(comanda).fetchall()
        name = []
        for i in result:
            if set(i[1:]) == {None} or set(i[:-1]):
                name.append(i[0])

        id = name[-1]
        self.table_of_classes.setItem(self.table_of_classes.rowCount() - 1, self.table_of_classes.columnCount() - 1,
                                      QTableWidgetItem(str(id)))
        con.commit()
        con.close()

    def update_loads(self, row, column):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        if self.g:
            if column == 4:
                ok_pressed = False
            if column == 2:
                subjects = cur.execute("SELECT Name FROM subjects_in_school").fetchall()
                value, ok_pressed = QInputDialog.getItem(
                    self, "Выбор предмета", "Выберите Предмет",
                    list(map(lambda x: x[0], subjects)), 1, False)
                table_sql = ("subjects_in_school", f"Name='{value}'")
                name_surname_middle_name = self.table_of_loads.cellWidget(row, 0).text().split(" ")
                cur.execute(
                    f"""UPDATE timetable SET id_subject = (SELECT id FROM subjects_in_school WHERE Name = '{value}') 
                WHERE id_subject in (SELECT id FROM subjects_in_school 
                WHERE Name = '{self.table_of_loads.cellWidget(row, 2).text()}') and 
                id_class in (SELECT id FROM list_of_classes 
                WHERE name_class = '{self.table_of_loads.cellWidget(row, 1).text()}') and 
                id_teacher in (SELECT id FROM list_of_teachers 
                WHERE Name = '{name_surname_middle_name[1]}' and 
                Surname = '{name_surname_middle_name[0]}' and 
                Middle_name = '{name_surname_middle_name[2]}')""").fetchall()
                time_lessons = cur.execute(f"""SELECT day_of_week, number_lesson, id_class FROM timetable 
                WHERE id_subject in (SELECT id FROM subjects_in_school 
                WHERE Name = '{value}') and 
                id_class in (SELECT id FROM list_of_classes 
                WHERE name_class = '{self.table_of_loads.cellWidget(row, 1).text()}') and 
                id_teacher in (SELECT id FROM list_of_teachers 
                WHERE Name = '{name_surname_middle_name[1]}' and 
                Surname = '{name_surname_middle_name[0]}' and 
                Middle_name = '{name_surname_middle_name[2]}')""").fetchall()
                classes = cur.execute(f"""SELECT * FROM list_of_classes""").fetchall()
                dic_of_nindex_column_class_in_timetable = {}
                for i in range(len(classes)):
                    dic_of_nindex_column_class_in_timetable[classes[i][1]] = i + 3
                for i in range(len(time_lessons)):
                    time_lessons[i] = list(time_lessons[i])

                    time_lessons[i][2] = cur.execute(f"""SELECT name_class FROM list_of_classes WHERE id = {time_lessons[i][2]}""").fetchall()[0][0]
                print(classes, dic_of_nindex_column_class_in_timetable)
                print(1)
                print(time_lessons)
                print(f"""SELECT day_of_week, number_lesson, id_class FROM timetable 
                WHERE id_subject in (SELECT id FROM subjects_in_school 
                WHERE Name = '{value}') and 
                id_class in (SELECT id FROM list_of_classes 
                WHERE name_class = '{self.table_of_loads.cellWidget(row, 1).text()}') and 
                id_teacher in (SELECT id FROM list_of_teachers 
                WHERE Name = '{name_surname_middle_name[1]}' and 
                Surname = '{name_surname_middle_name[0]}' and 
                Middle_name = '{name_surname_middle_name[2]}')""")
                for i in time_lessons:
                    print((i[0] - 1) * 6 + i[1] - 1, dic_of_nindex_column_class_in_timetable[i[2]])
                    self.table_of_timetable.cellWidget((i[0] - 1) * 6 + i[1] - 1, dic_of_nindex_column_class_in_timetable[i[2]]).setText(value)
            if column == 1:
                classes = cur.execute("SELECT * FROM list_of_classes").fetchall()
                value, ok_pressed = QInputDialog.getItem(
                    self, "Выбор класса", "Выберите класс",
                    list(map(lambda x: x[1:][0], classes)), 1, False)
                table_sql = ("list_of_classes", f"name_class = '{value}'")
            if column == 0:
                teachers = cur.execute("SELECT * FROM list_of_teachers").fetchall()
                try:
                    value, ok_pressed = QInputDialog.getItem(
                        self, "Выбор учителя", "Выберите учителя",
                        list(map(lambda x: " ".join([x[1:][1], x[1:][0], x[1:][2]]), teachers)), 1, False)
                except TypeError:
                    for i in range(len(teachers)):
                        teachers[i] = list(teachers[i])
                        for j in range(4):
                            if teachers[i][j] is None:
                                teachers[i][j] = ""
                    value, ok_pressed = QInputDialog.getItem(
                        self, "Выбор учителя", "Выберите учителя",
                        list(map(lambda x: " ".join([x[1:][1], x[1:][0], x[1:][2]]), teachers)), 1, False)
                t = value.split(" ")
                table_sql = ("list_of_teachers", f"Name = '{t[1]}' and Surname = '{t[0]}' and Middle_name = '{t[2]}'")
            if column == 3:
                value, ok_pressed = QInputDialog.getInt(
                    self, "Ввод нагрузки", "Введите нагрузку",
                    int(self.table_of_loads.cellWidget(row, column).text()), 1, 1000, 1)
                table_sql = "NO"
            if ok_pressed:
                self.table_of_loads.cellWidget(row, column).setText(str(value))
                name_of_column = cur.execute(f"""pragma table_info(list_of_loads)""").fetchall()
                if table_sql != "NO":
                    id_value = cur.execute(f"""SELECT id FROM {table_sql[0]} WHERE {table_sql[1]}""").fetchall()[0][0]
                else:
                    id_value = value
                cur.execute(f"""UPDATE list_of_loads SET {name_of_column[column][1]} = {id_value} 
                WHERE id = {self.table_of_loads.cellWidget(row, 4).text()}""")
            con.commit()
            con.close()
        self.g = True

    def update_timetable(self, row, column):
        if column not in (0, 1, 2):
            con = sqlite3.connect("BD/base_school.db")
            cur = con.cursor()
            subjects = cur.execute(
                f"SELECT Name FROM  subjects_in_school "
                f"WHERE id in (SELECT subject FROM list_of_loads "
                f"WHERE name_class in (SELECT id FROM list_of_classes "
                f"WHERE name_class = '{self.table_of_timetable.horizontalHeaderItem(column).text()}'))").fetchall()
            subject, ok_pressed = QInputDialog.getItem(
                self, "Выбор предмета", "Выберите Предмет",
                list(map(lambda x: x[0], subjects)) + [" "], 1, False)
            if ok_pressed:
                id_class = cur.execute(
                    f"""SELECT id FROM list_of_classes 
                    WHERE name_class = '{self.table_of_timetable.horizontalHeaderItem(column).text()}'""").fetchall()[
                    0][0]
                if subject != " ":
                    id_subject = \
                        cur.execute(f"""SELECT id FROM subjects_in_school WHERE Name = '{subject}'""").fetchall()[0][0]
                    id_teacher = cur.execute(f"""SELECT name_teacher FROM list_of_loads 
                    WHERE subject = {id_subject} and name_class = {id_class}""").fetchall()
                    self.table_of_timetable.cellWidget(row, column).setText(subject)
                    result = cur.execute(f"""SELECT * FROM timetable 
                    WHERE
                     number_lesson = {row % 6 + 1} and 
                    day_of_week = {row // 6 + 1} and 
                    id_class = {id_class}""").fetchall()
                    if result:
                        cur.execute(
                            f"""UPDATE timetable SET id_subject = {id_subject}, 
                            id_teacher = {id_teacher[0][0]}
                             WHERE number_lesson = {row % 6 + 1} and 
                    day_of_week = {row // 6 + 1} and 
                    id_class = {id_class}""")
                    else:
                        cur.execute(
                            f"""INSERT INTO timetable 
                            VALUES({id_class}, {row // 6 + 1}, {row % 6 + 1},
{id_subject}, {id_teacher[0][0]}, NULL)""")
                else:
                    self.table_of_timetable.cellWidget(row, column).setText(" ")
                    cur.execute(f"""DELETE FROM timetable WHERE number_lesson = {row % 6 + 1} and 
                        day_of_week = {row // 6 + 1} and 
                        id_class = {id_class}""")
                con.commit()
            con.close()

    def saving_name_of_school(self):
        file = open("txt files/settings.txt", "r").read().split("\n")
        file[0] = self.name_of_school.text()
        write_file = open("txt files/settings.txt", "w")
        write_file.write("\n".join(file))

    def saving__settings(self):
        list_ = [self.one_lesson.text(),
                  self.two_lesson.text(),
                  self.three_lesson.text(),
                  self.fout_lesson.text(),
                  self.five_lesson.text(),
                  self.six_lesson.text(),
                  self.seven_lesson.text(),
                  self.eight_lesson.text()]
        file = open("txt files/settings.txt", "r").read().split("\n")
        file[1:] = list_
        write_file = open("txt files/settings.txt", "w")
        write_file.write("\n".join(file))
        self.dic_time_management = {}
        for i in range(1, 7):
            self.dic_time_management[i] = file[i]
        self.dic_time_management[0] = file[6]
        self.updates_times_lessons()

    def updates_times_lessons(self):
        k = 1
        for i in range(36):
            if k > 6:
                k = 1
            self.table_of_timetable.setItem(i, 2, QTableWidgetItem(self.dic_time_management[k]))
            k += 1

    def call_update_teachers(self, value, value1):
        self.update_(value, value1, self.table_of_teachers, "list_of_teachers")

    def call_update_subjects(self, value, value1):
        self.update_(value, value1, self.table_of_subjects, "subjects_in_school")

    def call_update_classes(self, value, value1):
        self.update_(value, value1, self.table_of_classes, "list_of_classes")

    def update_(self, row, column, table_widget, sql_table):
        try:
            count_columns = table_widget.columnCount()
            id = table_widget.item(row, count_columns - 1).text()
            element = table_widget.item(row, column).text()
            index = column + 1
            con = sqlite3.connect("BD/base_school.db")
            cur = con.cursor()
            name_of_column = cur.execute(f"""pragma table_info({sql_table})""").fetchall()[index]
            if sql_table != "list_of_teachers":
                element_before = cur.execute(f"""SELECT {name_of_column[1]} FROM {sql_table} 
                WHERE id = {id}""").fetchall()[0][0]
            else:
                element_before = " ".join(list(map(lambda x: "" if not x else x, cur.execute(
                    f"""SELECT Surname, Name, Middle_name FROM list_of_teachers WHERE id = {id}""").fetchall()[0])))
            cur.execute(f"""UPDATE {sql_table} SET {name_of_column[1]} = '{element}'
            WHERE id = {id}""")
            if sql_table == "list_of_teachers":
                element = " ".join(list(map(lambda x: "" if not x else x, cur.execute(
                    f"""SELECT Surname, Name, Middle_name FROM list_of_teachers WHERE id = {id}""").fetchall()[0])))
            dic = {"list_of_classes": 1, "list_of_teachers": 0, "subjects_in_school": 2}
            column = dic[sql_table]
            for i in range(self.table_of_loads.rowCount()):
                if self.table_of_loads.cellWidget(i, column).text() == element_before:
                    self.table_of_loads.cellWidget(i, column).setText(element)
            con.commit()
            con.close()
            if sql_table == "list_of_classes":
                for i in range(3, self.table_of_timetable.columnCount()):
                    if self.table_of_timetable.horizontalHeaderItem(i).text() == element_before:
                        item = QTableWidgetItem(str(element))
                        item.setFont(QFont("Impact", 14))
                        self.table_of_timetable.setHorizontalHeaderItem(i, item)
        except IndexError:
            pass

    def delete(self, table_widget, table):
        if table_widget.rowCount():
            valid = QMessageBox.question(
                self, 'Удаление', 'Действительно хотите удалить этот элемент',
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                con = sqlite3.connect("BD/base_school.db")
                cur = con.cursor()
                current = table_widget.currentRow()
                dic = {"list_of_classes": 1, "list_of_teachers": 0, "subjects_in_school": 2}
                column = dic[table]
                name_of_column_in_loads = cur.execute(f"""pragma table_info(list_of_loads)""").fetchall()[column][
                    1]
                id = table_widget.item(current, table_widget.columnCount() - 1).text()
                result = cur.execute(
                    f"""SELECT * FROM list_of_loads WHERE {name_of_column_in_loads} = {id}""").fetchall()
                if result:
                    valid = QMessageBox.question(
                        self, '',
                        'Вы не можете удалить этот элемент из-за '
                        'того что он \nиспользуется в списке '
                        'нагрузок, чтобы его удалить\n '
                        'нужно удалит все нагрузки связанные\n с ним',
                        QMessageBox.Ok)
                else:
                    cur.execute(f"""DELETE FROM {table}
                    WHERE id = {table_widget.item(current, table_widget.columnCount() - 1).text()}""")
                    table_widget.removeRow(current)
                    con.commit()
                con.close()

    # Добавление Учителя
    def add(self, table_widget, table):
        con = sqlite3.connect("BD/base_school.db")
        cur = con.cursor()
        count_columns = table_widget.columnCount()
        table_widget.insertRow(table_widget.rowCount())
        cur.execute(f"""INSERT INTO {table} 
        VALUES ({("NULL, " * count_columns)[:-2]})""")
        name_of_column = cur.execute(f"""pragma table_info({table})""").fetchall()
        comanda = "SELECT * FROM " + table
        result = cur.execute(comanda).fetchall()
        name = []
        for i in result:
            if set(i[1:]) == {None} or set(i[:-1]):
                name.append(i[0])

        id = name[-1]
        item = QTableWidgetItem(str(id))
        table_widget.setItem(table_widget.rowCount() - 1, table_widget.columnCount() - 1, item)
        con.commit()
        con.close()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    form = MyWidget()
    form.show()
    sys.excepthook = except_hook
    sys.exit(app.exec())
